"""
Billing Software — Excel ledger backend.
JSON stays the source of truth; bills.xlsx is a derived export rewritten on every change.
"""
import os
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app import local_storage as ls
from app.local_storage import LocalStorage

logger = logging.getLogger("billing")

HEADER = ["Timestamp", "Bill No", "Customer", "Phone", "Items",
          "Total", "Paid", "Change", "Payment Type", "Status", "Deleted At", "Debit Account ID",
          "Print Requested", "Printed", "Subtotal", "Discount %", "Discount Amount"]


class ExcelStorage(LocalStorage):
    """LocalStorage + an Excel mirror at data/bills.xlsx."""

    def __init__(self, xlsx_path: str | None = None):
        self.xlsx_path = xlsx_path or os.path.join(ls.DATA_DIR, "bills.xlsx")

    # --- mutators: write JSON (super) then sync Excel ---
    def add_bill(self, *args, **kwargs):
        result = super().add_bill(*args, **kwargs)
        self._sync_excel()
        return result

    def edit_bill(self, *args, **kwargs):
        result = super().edit_bill(*args, **kwargs)
        self._sync_excel()
        return result

    def delete_bill(self, *args, **kwargs):
        result = super().delete_bill(*args, **kwargs)
        self._sync_excel()
        return result

    def set_setting(self, *args, **kwargs):
        result = super().set_setting(*args, **kwargs)
        self._sync_excel()
        return result

    def record_debit_payment(self, *args, **kwargs):
        result = super().record_debit_payment(*args, **kwargs)
        self._sync_excel()
        return result

    def mark_bill_print_status(self, *args, **kwargs):
        result = super().mark_bill_print_status(*args, **kwargs)
        self._sync_excel()
        return result

    def mark_debit_payment_print_status(self, *args, **kwargs):
        result = super().mark_debit_payment_print_status(*args, **kwargs)
        self._sync_excel()
        return result

    def _sync_excel(self):
        try:
            bills = ls._load_json(ls.BILLS_PATH, [])
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills"
            ws.append(HEADER)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"

            del_fill = PatternFill("solid", fgColor="FFF0F0")
            for b in bills:
                ws.append([
                    b.get("timestamp", ""), b.get("bill_no", ""), b.get("customer_name", ""),
                    b.get("phone", ""), b.get("items", ""), b.get("total", 0), b.get("paid", 0),
                    b.get("change", 0), b.get("payment_type", ""), b.get("status", "active"),
                    b.get("deleted_at", ""), b.get("account_id", ""),
                    b.get("print_requested"), b.get("printed"),
                    b.get("subtotal", b.get("total", 0)), b.get("discount_percent", 0),
                    b.get("discount_amount", 0),
                ])
                if b.get("status") == "deleted":
                    for cell in ws[ws.max_row]:
                        cell.fill = del_fill
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=8):
                for cell in row:
                    cell.number_format = u"₹#,##0"
            for row in ws.iter_rows(min_row=2, min_col=15, max_col=17):
                row[0].number_format = u"₹#,##0.00"
                row[1].number_format = "0.00%"
                row[1].value = float(row[1].value or 0) / 100
                row[2].number_format = u"₹#,##0.00"

            summary = wb.create_sheet("Summary")
            summary.append(["Date", "Cash Collected", "UPI Collected", "Debit Sales",
                            "Total Sales", "Total Collected", "Bill Count"])
            for cell in summary[1]:
                cell.font = Font(bold=True)
            days: dict[str, dict] = {}
            for b in bills:
                if b.get("status") == "deleted":
                    continue
                day = str(b.get("timestamp", ""))[:10]
                if not day:
                    continue
                agg = days.setdefault(day, {"Cash": 0.0, "UPI": 0.0, "Debit": 0.0,
                                            "sales": 0.0, "collected": 0.0, "count": 0})
                amount = float(b.get("total", 0) or 0)
                paid = float(b.get("paid", amount) or 0)
                agg["sales"] += amount
                agg["collected"] += paid
                agg["count"] += 1
                ptype = b.get("payment_type", "Cash")
                if ptype in ("Cash", "UPI"):
                    agg[ptype] += paid
                elif ptype == "Debit":
                    agg["Debit"] += amount

            payments = ls._load_json(ls.DEBIT_PAYMENTS_PATH, [])
            for payment in payments:
                if payment.get("status", "active") != "active":
                    continue
                day = str(payment.get("timestamp", ""))[:10]
                if not day:
                    continue
                agg = days.setdefault(day, {"Cash": 0.0, "UPI": 0.0, "Debit": 0.0,
                                            "sales": 0.0, "collected": 0.0, "count": 0})
                amount = float(payment.get("amount", 0) or 0)
                method = payment.get("payment_method", "")
                if method in ("Cash", "UPI"):
                    agg[method] += amount
                agg["collected"] += amount
            for day in sorted(days):
                a = days[day]
                summary.append([day, a["Cash"], a["UPI"], a["Debit"],
                                a["sales"], a["collected"], a["count"]])

            ledger = wb.create_sheet("Debit Ledger")
            ledger.append(["Timestamp", "Account ID", "Customer", "Phone", "Entry Type",
                           "Reference", "Items / Note", "Debit", "Credit", "Method", "Printed",
                           "Subtotal", "Discount %", "Discount Amount"])
            for cell in ledger[1]:
                cell.font = Font(bold=True)
            ledger.freeze_panes = "A2"
            for b in bills:
                if b.get("status") == "deleted" or str(b.get("payment_type", "")).lower() != "debit":
                    continue
                ledger.append([
                    b.get("timestamp", ""), b.get("account_id", ""), b.get("customer_name", ""),
                    b.get("phone", ""), "Purchase", f"Bill #{b.get('bill_no', '')}",
                    b.get("items", ""), b.get("total", 0), 0, "Debit", b.get("printed"),
                    b.get("subtotal", b.get("total", 0)), b.get("discount_percent", 0),
                    b.get("discount_amount", 0),
                ])
            for p in payments:
                if p.get("status", "active") != "active":
                    continue
                ledger.append([
                    p.get("timestamp", ""), p.get("account_id", ""), p.get("customer_name", ""),
                    p.get("phone", ""), "Payment", p.get("payment_id", "")[:8],
                    p.get("note", ""), 0, p.get("amount", 0), p.get("payment_method", ""), p.get("printed"),
                    0, 0, 0,
                ])
            for row in ledger.iter_rows(min_row=2, min_col=8, max_col=9):
                for cell in row:
                    cell.number_format = u"₹#,##0.00"

            accounts = wb.create_sheet("Debit Accounts")
            accounts.append(["Account ID", "Customer", "Phone", "Total Purchased",
                             "Total Paid", "Outstanding", "Status", "Last Activity"])
            for cell in accounts[1]:
                cell.font = Font(bold=True)
            for account in self.get_debit_accounts():
                accounts.append([
                    account["account_id"], account["customer_name"], account["phone"],
                    account["total_purchased"], account["total_paid"], account["balance"],
                    account["status"], account["last_activity"],
                ])

            tmp = self.xlsx_path + ".tmp"
            wb.save(tmp)
            os.replace(tmp, self.xlsx_path)
        except (PermissionError, OSError) as e:
            logger.warning(f"Excel sync skipped (file open/locked?): {e}")
