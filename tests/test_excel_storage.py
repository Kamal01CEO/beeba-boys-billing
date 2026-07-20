import os
import json
import pytest
from openpyxl import load_workbook
import app.local_storage as ls


@pytest.fixture
def storage(tmp_path, monkeypatch):
    monkeypatch.setattr(ls, "DATA_DIR", str(tmp_path))
    monkeypatch.setattr(ls, "BILLS_PATH", str(tmp_path / "bills.json"))
    monkeypatch.setattr(ls, "SETTINGS_PATH", str(tmp_path / "settings.json"))
    monkeypatch.setattr(ls, "DEBIT_PAYMENTS_PATH", str(tmp_path / "debit_payments.json"))
    from app.excel_storage import ExcelStorage
    return ExcelStorage(xlsx_path=str(tmp_path / "bills.xlsx"))


def test_add_bill_writes_json_and_xlsx(storage, tmp_path):
    storage.add_bill("Ramesh", "911", [{"name": "Shirt", "qty": 2, "price": 800}], 1600, 1600, "Cash")
    # JSON is the source of truth
    bills = json.load(open(str(tmp_path / "bills.json")))
    assert bills[-1]["customer_name"] == "Ramesh"
    # xlsx mirrors it
    wb = load_workbook(str(tmp_path / "bills.xlsx"))
    assert "Bills" in wb.sheetnames and "Summary" in wb.sheetnames
    rows = list(wb["Bills"].iter_rows(values_only=True))
    assert rows[0][2] == "Customer"          # header
    assert rows[1][2] == "Ramesh"            # data row


def test_discount_breakdown_is_exported(storage, tmp_path):
    storage.add_bill(
        "Ramesh", "911", [{"name": "Shirt", "qty": 1, "price": 1000}],
        900, 900, "Cash", subtotal=1000, discount_percent=10, discount_amount=100,
    )
    wb = load_workbook(str(tmp_path / "bills.xlsx"), data_only=True)
    rows = list(wb["Bills"].iter_rows(values_only=True))
    assert rows[0][14:17] == ("Subtotal", "Discount %", "Discount Amount")
    assert rows[1][14] == 1000
    assert rows[1][15] == pytest.approx(0.10)
    assert rows[1][16] == 100


def test_summary_totals(storage, tmp_path):
    storage.add_bill("A", "", [{"name": "Shirt", "qty": 1, "price": 100}], 100, 100, "Cash")
    storage.add_bill("B", "", [{"name": "Jeans", "qty": 1, "price": 200}], 200, 200, "UPI")
    b3 = storage.add_bill("C", "", [{"name": "T", "qty": 1, "price": 50}], 50, 50, "Cash")
    storage.delete_bill(b3)  # excluded from summary
    wb = load_workbook(str(tmp_path / "bills.xlsx"))
    summary = list(wb["Summary"].iter_rows(values_only=True))
    assert summary[0] == ("Date", "Cash Collected", "UPI Collected", "Debit Sales",
                          "Total Sales", "Total Collected", "Bill Count")
    # one day row: Cash=100, UPI=200, Total=300, Count=2
    assert summary[1][1] == 100 and summary[1][2] == 200
    assert summary[1][4] == 300 and summary[1][5] == 300 and summary[1][6] == 2


def test_debit_sheets_show_account_and_ledger(storage, tmp_path):
    purchase = storage.create_debit_purchase(
        "Kamal", "911", [{"name": "Shirt", "qty": 1, "price": 500}], 500
    )
    storage._sync_excel()
    storage.record_debit_payment(purchase["account_id"], 200, "Cash")
    wb = load_workbook(str(tmp_path / "bills.xlsx"))
    assert "Debit Accounts" in wb.sheetnames
    assert "Debit Ledger" in wb.sheetnames
    account_rows = list(wb["Debit Accounts"].iter_rows(values_only=True))
    assert account_rows[1][1] == "Kamal"
    assert account_rows[1][5] == 300
    ledger_rows = list(wb["Debit Ledger"].iter_rows(values_only=True))
    assert [row[4] for row in ledger_rows[1:]] == ["Purchase", "Payment"]


def test_lock_guard_never_loses_bill(storage, tmp_path, monkeypatch):
    from openpyxl import Workbook
    monkeypatch.setattr(Workbook, "save", lambda self, *a, **k: (_ for _ in ()).throw(PermissionError("open in Excel")))
    # Must not raise; JSON must still record the bill
    bill_no = storage.add_bill("Locked", "", [{"name": "Shirt", "qty": 1, "price": 500}], 500, 500, "Cash")
    assert bill_no >= 1
    bills = json.load(open(str(tmp_path / "bills.json")))
    assert bills[-1]["customer_name"] == "Locked"
